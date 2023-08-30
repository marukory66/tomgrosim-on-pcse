#%%
import pandas as pd
import re
from datetime import datetime
from datetime import date as dt
import openpyxl as xl

# chamber_explantory = "C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/plantation_data/tomato_chamber_explanatory.csv"
# df_chamber_explantory = pd.read_csv(chamber_explantory)

# nl1 = "C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse/test_data/nl1.xls"
# df_nl1 = pd.read_excel(nl1)


def temperature_outside_chamber(chamber_explantory,nl1):
    df_chamber_explantory = pd.read_csv(chamber_explantory)
    date = df_chamber_explantory["Date"]
    IRRAD = df_chamber_explantory["Solar.radiation.accu"]
    TMIN = df_chamber_explantory["Temperature.outside.chamber.nightave"]
    TMAX = df_chamber_explantory["Temperature.outside.chamber.dayave"]
    for i in range(len(date)):
        tmp = date[i]
        # tmp = datetime(tmp,'%Y-%m-%d')
        # date[i] = date[i].replace("-","/")
        # date[i] = date[i].replace("/0","/")
        date[i] = pd.to_datetime(date[i],format='%Y/%m/%d')
        # y = date[i].split("/")[0]
        # m = date[i].split("/")[1]
        # d = date[i].split("/")[2]
        # date[i] = datetime(int(y),int(m),int(d))
        # date[i].value = datetime(int(y),int(m),int(d))
        date[i] = date[i].to_pydatetime()
    #chamber_explantoryのデータに差し替える

    df_nl1 = pd.read_excel(nl1)

    for i in range(len(date)):
        df_nl1.iloc[11+i,0] = date[i]
        df_nl1.iloc[11+i,1] = IRRAD[i]
        df_nl1.iloc[11+i,2] = TMIN[i]
        df_nl1.iloc[11+i,3] = TMAX[i]
    return df_nl1


# df_nl1 = temperature_outside_chamber(df_chamber_explantory,df_nl1)
# df_nl1.to_excel('C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/df.xls',index = False)


#%%
# import openpyxl as xl


# wb = xl.load_workbook("C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/df.xlsx")
# sheet = ws["Sheet1"]
# for i in range(306):
#     n = 11+i
#     sheet.cell(row=n,columns=0).number_format = xl.styles.numbers.FORMAT_DATE_YYMMDD


# df_nl1.to_csv('C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/df.csv',index = False)
# wb = openpyxl.load_workbook("C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/df.xlsx")

# print(wb.sheetnames)"Sheet1"


# for i in range(len(date)):
#     y = date[i].split("-")[0]
#     m = date[i].split("-")[1]
#     d = date[i].split("-")[2]
#     cell = wb[11+i,0]
#     cell.value = datetime(int(y),int(m),int(d))


#無駄な部分を消す　不完全　要確認
# df_nl1 = df_nl1.drop([11+len(date),len(df_nl1)-1],axis = 1)

# with pd.ExcelWriter("C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/df.xls", datetime_format="YYYY-MM-DD HH:MM:SS") as writer:
#     df_nl1.to_excel(writer,index = False)


# df_nl1.to_excel('C:/Users/maruko/OneDrive - 愛媛大学 (1)/02_PCSE/tomgrosim-on-pcse_pypl/630nl1.xls',index = False)

# %%
